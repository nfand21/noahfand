from flask import Flask, request, render_template, send_file
import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
import os

app = Flask(__name__)

def get_otc_stock_data_yf(ticker):
    metrics = {
        'Ticker': ticker,
        'Company Full Name': "N/A",
        'Country of Incorporation': "N/A",
        'Last Price': "N/A",
        'Market Cap': "N/A",
        'Enterprise Value': "N/A",
        'Last Year Revenue': "N/A",
        'EV / Sales': "N/A",
        'Gross Margin': "N/A",
        'P/E': "N/A"
    }

    try:
        stock = yf.Ticker(ticker.strip())
        info = stock.info
        financials = stock.financials

        # General Info
        metrics['Company Full Name'] = info.get('longName', "N/A")
        metrics['Country of Incorporation'] = info.get('country', "N/A")

        # Financial Metrics
        metrics['Last Price'] = info.get('previousClose', "N/A")
        metrics['Market Cap'] = info.get('marketCap', "N/A")
        metrics['Enterprise Value'] = info.get('enterpriseValue', "N/A")

        # Revenue Last Year
        try:
            revenue_last_year = financials.loc["Total Revenue"].iloc[0]
            metrics['Last Year Revenue'] = revenue_last_year
        except KeyError:
            metrics['Last Year Revenue'] = "N/A"

        # EV / Sales
        if metrics['Enterprise Value'] != "N/A" and metrics['Last Year Revenue'] != "N/A":
            metrics['EV / Sales'] = metrics['Enterprise Value'] / metrics['Last Year Revenue']
        else:
            metrics['EV / Sales'] = "N/A"

        # Gross Margin
        gross_margin = info.get('grossMargins', "N/A")
        if gross_margin != "N/A":
            metrics['Gross Margin'] = gross_margin * 100

        # P/E Ratio
        metrics['P/E'] = info.get('trailingPE', "N/A")

    except Exception as e:
        metrics['Error'] = f"Could not retrieve data: {e}"

    return metrics

def autofit_columns_and_format(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Auto-fit columns
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column_letter].width = max_length + 2

    # Apply number and percent formatting
    currency_style = NamedStyle(name="comma_format", number_format="#,##0")
    percent_style = NamedStyle(name="percent_format", number_format="0.00%")
    for row in sheet.iter_rows(min_row=2):  # Skip header row
        # Market Cap, Enterprise Value, Last Year Revenue (comma format)
        for col in [5, 6, 7]:  # Columns E, F, G
            cell = row[col - 1]
            if isinstance(cell.value, (int, float)):
                cell.style = currency_style

        # Gross Margin (percent format)
        cell = row[8 - 1]  # Column I
        if isinstance(cell.value, (int, float)):
            cell.style = percent_style

    workbook.save(filename)

def get_next_filename(base_name="Stock Data", ext=".xlsx"):
    """Generates the next filename in the sequence (e.g., Stock Data 1, Stock Data 2, etc.)."""
    i = 1
    while os.path.exists(f"{base_name} {i}{ext}"):
        i += 1
    return f"{base_name} {i}{ext}"

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        tickers = request.form["tickers"].upper()
        ticker_list = tickers.split(",")

        all_data = []
        for ticker in ticker_list:
            data = get_otc_stock_data_yf(ticker)
            all_data.append(data)

        # Create a DataFrame
        df = pd.DataFrame(all_data)

        # Specify column order
        column_order = [
            'Ticker', 'Company Full Name', 'Country of Incorporation',
            'Last Price', 'Market Cap', 'Enterprise Value',
            'Last Year Revenue', 'EV / Sales', 'Gross Margin', 'P/E'
        ]
        df = df[column_order]

        # Generate dynamic file name
        output_file = get_next_filename()
        df.to_excel(output_file, index=False)

        # Auto-fit columns and apply formatting
        autofit_columns_and_format(output_file)

        return send_file(output_file, as_attachment=True)

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
